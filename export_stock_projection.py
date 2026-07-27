"""
Export Excel — Projection de stock CEGECLIM
Appelé depuis l'API Next.js : reçoit les données en JSON sur stdin.
"""
import sys, json, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data_in = json.loads(sys.stdin.read())
rows       = data_in.get("rows", [])
FAMILLE    = data_in.get("famille", "")
MACRO      = data_in.get("macro", "")
GRANULARITE= data_in.get("granularite", "mensuel")  # hebdo | mensuel
OUT        = data_in.get("out", "/tmp/projection_stock.xlsx")

if not rows:
    print("Aucune donnée.", file=sys.stderr); sys.exit(0)

# ────────────────────────────────────────────────────────────────────────────
# Agrégation mensuelle
# ────────────────────────────────────────────────────────────────────────────
def mois(d): return d[:7]

def agr_mensuel(rows):
    sum_fields = ["commandes_fournisseurs_attendues","besoins_clients_fermes",
                  "besoins_clients_retard","prevision_base_n1_origine",
                  "prevision_base_n1","prevision_ventes","prevision_transferee_entrante"]
    key_fields = ["reference_article","designation","famille","macro_famille",
                  "fournisseur_principal","depot","statut_substitution"]
    poids = {"ROUGE":3,"ORANGE":2,"JAUNE":1,"VERT":0}
    agg = {}
    for r in rows:
        k = tuple(r.get(f,"") for f in key_fields) + (mois(r["periode_debut"]),)
        if k not in agg:
            agg[k] = {f: r.get(f,"") for f in key_fields}
            agg[k]["periode_debut"]  = mois(r["periode_debut"])
            agg[k]["stock_initial"]  = r.get("stock_initial") or 0
            agg[k]["stock_securite"] = r.get("stock_securite") or 0
            agg[k]["stock_projete"]  = r.get("stock_projete") or 0
            agg[k]["niveau_alerte"]  = r.get("niveau_alerte","VERT")
            agg[k]["date_rupture"]   = r.get("date_rupture") or ""
            agg[k]["coefficient_prevision_applique"] = r.get("coefficient_prevision_applique") or 0
            for f in sum_fields: agg[k][f] = 0
        for f in sum_fields:
            agg[k][f] = (agg[k][f] or 0) + (r.get(f) or 0)
        agg[k]["stock_projete"] = r.get("stock_projete") or 0
        if (poids.get(r.get("niveau_alerte","VERT"),0) > poids.get(agg[k]["niveau_alerte"],0)):
            agg[k]["niveau_alerte"] = r.get("niveau_alerte","VERT")
        rd = r.get("date_rupture") or ""
        if rd and (not agg[k]["date_rupture"] or rd < agg[k]["date_rupture"]):
            agg[k]["date_rupture"] = rd
    return list(agg.values())

if GRANULARITE == "mensuel":
    rows = agr_mensuel(rows)

all_periods = sorted(set(r["periode_debut"] for r in rows))

# ────────────────────────────────────────────────────────────────────────────
# Couleurs
# ────────────────────────────────────────────────────────────────────────────
C = {
    "fond":"0B1220","creme":"F5F3EC","sauge":"A6A181","blanc":"FFFFFF",
    "violet":"7A5EA8","orange_al":"C1683C","vert":"3F9142","bleu":"4B92AC",
    "gris":"8A93A6","fond_ref":"F0EDE4","fond_n1":"E8F0E9","fond_prev":"FFF3E0",
    "fond_cdc":"F3EEF8","fond_cf":"E3F0F4","fond_stk":"E8F4FD",
    "rouge_al":"C1683C","jaune_al":"D69A4A","vert_al":"4B92AC",
}
def fill(h): return PatternFill("solid",fgColor=h)
def font(bold=False,color="141A26",size=9,name="Arial"):
    return Font(name=name,bold=bold,color=color,size=size)
def thin(): s=Side(border_style="thin",color="D0CAC0"); return Border(left=s,right=s,top=s,bottom=s)

GROUPES = [
    ("BL N-1",     ["prevision_base_n1_origine"],                        C["fond_n1"],   "3F9142"),
    ("PRÉVISIONS", ["coefficient_prevision_applique","prevision_ventes"], C["fond_prev"], "C1683C"),
    ("CDC FERMES", ["besoins_clients_fermes","besoins_clients_retard"],   C["fond_cdc"],  "7A5EA8"),
    ("ENTRÉES CF", ["commandes_fournisseurs_attendues"],                  C["fond_cf"],   "4B92AC"),
    ("STOCK",      ["stock_projete","stock_securite"],                    C["fond_stk"],  "0B1220"),
]
LABELS = {
    "prevision_base_n1_origine":"Ventes N-1","coefficient_prevision_applique":"Hyp. ×",
    "prevision_ventes":"Prévision","besoins_clients_fermes":"CDC fermes",
    "besoins_clients_retard":"CDC retard","commandes_fournisseurs_attendues":"Entrées CF",
    "stock_projete":"Stock projeté","stock_securite":"Stk sécurité",
}
metrics   = [m for _,ml,_,_ in GROUPES for m in ml]
n_metrics = len(metrics)
n_periods = len(all_periods)
INFO_COLS = ["macro_famille","famille","reference_article","designation",
             "fournisseur_principal","stock_initial","statut_substitution",
             "niveau_alerte","date_rupture"]
INFO_LBL  = ["Fam. macro","Famille","Référence","Désignation",
             "Fournisseur","Stk actuel","Statut","Alerte","Rupture"]

# ────────────────────────────────────────────────────────────────────────────
# Workbook
# ────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Légende ─────────────────────────────────────────────────────────────────
leg = wb.active; leg.title="Légende"; leg.sheet_view.showGridLines=False
leg.column_dimensions["A"].width=3; leg.column_dimensions["B"].width=30; leg.column_dimensions["C"].width=55
rows_leg = [
    ("PROJECTION DE STOCK — CEGECLIM",None,None),
    (None,None,None),
    ("COULEURS",None,None),
    (None,"Fond crème","Identité référence (famille, désignation, fournisseur)"),
    (None,"Fond vert pâle","Ventes N-1"),
    (None,"Fond orange pâle","Prévisions (hypothèse × et volume)"),
    (None,"Fond violet pâle","CDC fermes et retards"),
    (None,"Fond bleu pâle","Entrées fournisseur attendues"),
    (None,"Fond bleu clair","Projection du stock"),
    (None,None,None),
    ("ALERTES",None,None),
    (None,"ROUGE","Stock insuffisant"),
    (None,"ORANGE","Stock tendu"),
    (None,"JAUNE","À surveiller"),
    (None,"VERT","Stock satisfaisant"),
    (None,None,None),
    ("REMPLACEMENTS",None,None),
    (None,"REMPLACEE","Besoins transférés — prévision 0"),
    (None,"REMPLACANTE","Reprend l'historique d'une ou plusieurs références"),
    (None,"PARTIELLE","Transfert partiel"),
    (None,"ACTIVE","Référence courante"),
    (None,None,None),
    ("PÉRIMÈTRE",None,None),
    (None,"Famille macro",MACRO or "(toutes)"),
    (None,"Famille",FAMILLE or "(toutes)"),
    (None,"Granularité",GRANULARITE.capitalize()),
    (None,"Généré le",datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
]
for i,(a,b,c) in enumerate(rows_leg,1):
    if a and not b:
        ce=leg.cell(i,2,a); ce.font=Font(name="Arial",bold=True,size=10,color=C["blanc"]); ce.fill=fill(C["sauge"])
    else:
        if b: leg.cell(i,2,b).font=font(bold=True)
        if c: leg.cell(i,3,c).font=font()

# ── Projection ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Projection stock")
ws.sheet_view.showGridLines = False
ws.freeze_panes = f"{get_column_letter(len(INFO_COLS)+1)}4"

# Titre ligne 1
titre = (FAMILLE or MACRO or "Toutes familles") + " — Projection stock " + GRANULARITE.upper()
c=ws.cell(1,1,titre); c.font=Font(name="Arial",bold=True,size=13,color=C["blanc"]); c.fill=fill(C["fond"])
c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=26
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(INFO_COLS)+n_metrics*n_periods)

grp_start = len(INFO_COLS)+1

# Infos fixes : fusion lignes 2-3
for ci,lbl in enumerate(INFO_LBL,1):
    ws.merge_cells(start_row=2,start_column=ci,end_row=3,end_column=ci)
    c=ws.cell(2,ci,lbl)
    c.font=Font(name="Arial",bold=True,size=8,color=C["blanc"]); c.fill=fill(C["fond"])
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin()

# En-tête ligne 2 : période (fusionnée sur les métriques du groupe)
for pi,pdate in enumerate(all_periods):
    start_c = grp_start + pi*n_metrics
    end_c   = start_c + n_metrics - 1
    if GRANULARITE=="hebdo":
        try: lp=datetime.date.fromisoformat(pdate).strftime("S%V\n%d/%m")
        except: lp=pdate
    else:
        try:
            d=datetime.date.fromisoformat(pdate+"-01")
            lp=d.strftime("%b %Y")
        except: lp=pdate
    ws.merge_cells(start_row=2,start_column=start_c,end_row=2,end_column=end_c)
    c=ws.cell(2,start_c,lp); c.font=Font(name="Arial",bold=True,size=7,color=C["blanc"])
    c.fill=fill(C["sauge"]); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin()

# En-tête ligne 3 : métriques
for pi in range(n_periods):
    for gi,(grp,ml,fond_g,col_g) in enumerate(GROUPES):
        for mi,metric in enumerate(ml):
            off = sum(len(m) for _,m,_,_ in GROUPES[:gi])
            ac = grp_start + pi*n_metrics + off + mi
            c=ws.cell(3,ac,LABELS.get(metric,metric))
            c.font=Font(name="Arial",bold=True,size=7,color=col_g); c.fill=fill(fond_g)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin()
ws.row_dimensions[2].height=26; ws.row_dimensions[3].height=26

# Index données
idx={}
for r in rows: idx[(r["reference_article"],r["periode_debut"])]=r
refs_order=[]; refs_seen={}
for r in rows:
    ref=r["reference_article"]
    if ref not in refs_seen: refs_seen[ref]=r; refs_order.append(ref)

AL_FOND={"ROUGE":"FEE2D5","ORANGE":"FEF3CD","JAUNE":"FEFBCA","VERT":"D1E7DD"}
ST_COLOR={"REMPLACEE":C["gris"],"REMPLACANTE":C["vert"],"PARTIELLE":C["jaune_al"],"ACTIVE":C["fond"]}

row_num=4
for ref in refs_order:
    r0=refs_seen[ref]; statut=r0.get("statut_substitution") or "ACTIVE"
    row_fond="F8F5EF" if statut=="REMPLACEE" else "FFFFFF"
    info_vals=[r0.get("macro_famille",""),r0.get("famille",""),r0.get("reference_article",""),
               r0.get("designation",""),r0.get("fournisseur_principal","") or "",
               r0.get("stock_initial") or 0,statut,r0.get("niveau_alerte","VERT"),r0.get("date_rupture") or ""]
    for ci,val in enumerate(info_vals,1):
        c=ws.cell(row_num,ci,val); c.fill=fill(row_fond); c.border=thin()
        c.font=font(size=8); c.alignment=Alignment(horizontal="left",vertical="center")
        if ci==3: c.font=Font(name="Arial",bold=True,size=8,color=ST_COLOR.get(statut,C["fond"]))
        if ci==8:
            al=str(val); c.fill=fill(AL_FOND.get(al,row_fond))
            c.font=Font(name="Arial",bold=True,size=7,
                        color={"ROUGE":C["rouge_al"],"ORANGE":C["jaune_al"],"VERT":C["vert_al"]}.get(al,C["fond"]))
            c.alignment=Alignment(horizontal="center",vertical="center")
        if ci==9 and val: c.font=Font(name="Arial",size=8,color=C["rouge_al"])

    for pi,pdate in enumerate(all_periods):
        r=idx.get((ref,pdate),{})
        for gi,(grp,ml,fond_g,col_g) in enumerate(GROUPES):
            for mi,metric in enumerate(ml):
                off=sum(len(m) for _,m,_,_ in GROUPES[:gi])
                ac=grp_start+pi*n_metrics+off+mi
                val=r.get(metric)
                if val is None: val=""
                elif metric=="coefficient_prevision_applique": val=round(float(val),2)
                else: val=round(float(val),1) if val else 0
                c=ws.cell(row_num,ac,val); c.fill=fill(fond_g); c.border=thin()
                c.font=Font(name="Arial",size=8,color=C["fond"])
                c.alignment=Alignment(horizontal="right",vertical="center")
                if metric=="coefficient_prevision_applique": c.number_format='0.00"×"'
                elif metric=="stock_projete":
                    if isinstance(val,(int,float)) and val<0:
                        c.font=Font(name="Arial",bold=True,size=8,color=C["rouge_al"])
                    c.number_format="#,##0;[Red]-#,##0;-"
                elif metric not in ("","coefficient_prevision_applique"):
                    c.number_format="#,##0.#"

    ws.row_dimensions[row_num].height=15
    row_num+=1

# Largeurs
for ci,w in enumerate([14,14,18,42,18,9,14,8,11],1):
    ws.column_dimensions[get_column_letter(ci)].width=w
for pi in range(n_periods):
    for mi in range(n_metrics):
        ws.column_dimensions[get_column_letter(grp_start+pi*n_metrics+mi)].width=8.5

ws.auto_filter.ref=f"A3:{get_column_letter(len(INFO_COLS))}3"
wb.save(OUT)
print(f"OK|{len(refs_order)}|{n_periods}", flush=True)
